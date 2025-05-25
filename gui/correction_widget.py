import sys
import os
import pathlib
import pandas as pd
import openpyxl as xl
from PyQt6.QtWidgets import (
    QFileDialog, QWidget, QMessageBox, QTableWidgetItem
)
from PyQt6.QtCore import Qt, QEvent
from PyQt6.QtGui import QMouseEvent
from PyQt6 import uic
from core.processor import evaluate_dataframe
from gui.helpers import DataFrameViewer

class CorrectionWindow(QWidget):
    """
    A window for correcting and evaluating text responses from an Excel file.

    Provides functionality to:
    - Display and edit text responses and their corrections
    - Apply corrections to the original Excel file
    - Evaluate responses using AI
    - Filter and manage response data

    Attributes:
        df (pd.DataFrame): DataFrame containing responses and corrections
        df_evaluated (pd.DataFrame): Subset of df with evaluated responses
        totalTexts (int): Total number of texts to correct
        relations (dict): Mapping between processed and original texts
        idxTexts (int): Column index of text responses in original file
        idxCorr (int): Column index of corrections in original file
        file_name (str): Path to the original Excel file
        sheet_name (str): Name of the worksheet being processed
        wb (openpyxl.Workbook): Loaded Excel workbook
        excel (openpyxl.Worksheet): Active worksheet
    """
    def __init__(self, df, relations, file_name, idxTexts, idxCorr, numTotalOriginal, sheet_name):
        """
        Initialize the correction window with data and UI.

        Args:
            df: Processed DataFrame containing responses
            relations: Mapping between processed and original texts
            file_name: Path to the original Excel file
            idxTexts: Column index of text responses in original file
            idxCorr: Column index of corrections in original file
            numTotalOriginal: Total number of original responses
            sheet_name: Name of the worksheet being processed
        """
        super().__init__()

        # Initialize data attributes
        self.df = df
        self.df_evaluated = self.df[self.df['Nota'] != '']
        self.totalTexts = len(self.df)
        self.df_copy = df.copy()
        self.relations = relations
        self.idxTexts = idxTexts
        self.idxCorr = idxCorr
        self.file_name = file_name
        self.path = pathlib.PurePath(self.file_name)
        self.sheet_name = sheet_name

        # Load Excel workbook
        self.wb = xl.load_workbook(self.file_name)
        self.excel = self.wb[self.sheet_name]

        # Setup UI
        uic.loadUi('assets/CorrectionWindow.ui', self)
        self.progressBar.setRange(0,len(self.df['Nota']))
        self.progressBar.setValue(0)
        self.label_redu.setText("De " + str(numTotalOriginal) + " respuestas se ha reducido a " + str(len(self.df)))

        # Setup dialog and event filters
        self.dlg = QMessageBox(self)
        self.dlg.setWindowTitle("Error")
        self.table.viewport().installEventFilter(self)
        self.table.setContextMenuPolicy(Qt.ContextMenuPolicy.NoContextMenu)

        # Connect signals
        self.aplicarCorr.clicked.connect(self.make_corrections)
        self.cerrarGuardar.clicked.connect(self._close_and_save)
        self.testLLM.clicked.connect(self.test_LLM)
        self.corregir_IA.clicked.connect(self.evaluate)
        self.hideEval.stateChanged.connect(self.hide_evaluated)

        # Initialize table
        self.write_table()
        self.show()
    
    def write_table(self, remove=False, resize=True):
        """
        Populate the table widget with DataFrame contents.

        Args:
            remove: If True, removes already evaluated rows from display
        """
        
        self.df.fillna('', inplace=True)
        self.df_evaluated = self.df[self.df['Nota'] != '']

        if remove:
            self.df = self.df[self.df['Nota'] == '']

        # Configure table dimensions and headers
        self.table.setRowCount(self.df.shape[0])
        self.table.setColumnCount(self.df.shape[1])
        self.table.setHorizontalHeaderLabels(self.df.columns)

        # Populate table cells
        for i in range(len(self.df)):
            for j in range(len(self.df.columns)):
                if j == 2:
                    item = QTableWidgetItem()
                    item.setData(Qt.ItemDataRole.DisplayRole, int(self.df.iat[i,j]))
                    self.table.setItem(i,j, item)
                else:
                    self.table.setItem(i,j, QTableWidgetItem(str(self.df.iat[i,j])))
        
        # Configure column widths
        horizontalHeader = self.table.horizontalHeader()
        horizontalHeader.resizeSection(0, 640)
        horizontalHeader.resizeSection(1, 50)
        horizontalHeader.resizeSection(2, 50)
        horizontalHeader.resizeSection(3, 50)

        if resize:
            self.table.resizeRowsToContents()

        return

    
    def make_corrections(self):
        """Apply manual corrections to the original Excel file."""
        for row in range(self.table.rowCount()):

            score = self.table.item(row, 1).text()
            text = self.table.item(row, 0).text()
            
            if text not in self.relations:
                self.dlg.setText("Error en el texto: ", text, " -> ESTE ERROR NO DEBERIA PASSAR")
                button = self.dlg.exec()
                if button == QMessageBox.StandardButton.Ok:
                    return

            # Write corrections to Excel file
            if score != '':
                for idRow in self.relations[text]:
                    cell_to_write = self.excel.cell(row=idRow + 2, column=self.idxCorr + 1)
                    cell_to_write.value = str(score)

        # Update progress display
        self.df = self._dataframe_generation_from_table(self.table)
        evaluated_count = self.df['Nota'].apply(lambda x: pd.notna(x) and str(x) != '').sum()
        self.progressBar.setValue(evaluated_count)
        self.label_prog.setText(str(evaluated_count) + "/" + str(self.totalTexts) + ' Respuestas corregidas')

        self.write_table(remove=False)

        return
    
    def test_LLM(self):
        """Test the AI evaluation on already evaluated responses."""
        df_result = evaluate_dataframe(self.df_evaluated, test=True)
        df_result = df_result[df_result['nota IA'].notna()].reset_index()
        
        columns_to_show = ['Respuesta', 'Nota', 'nota IA', 'feedback IA', 'confidence']
        df_result = df_result[columns_to_show]

        # Show comparison in a new table
        if not df_result.empty:
            viewer = DataFrameViewer(df_result, parent=self)
            viewer.exec()

        return
    
    def hide_evaluated(self):
        """Toggle visibility of already evaluated rows based on checkbox state."""
        column_to_check = 1 # Grage Colum
        hide_non_empty = self.hideEval.isChecked()

        for row in range(self.table.rowCount()):
            item = self.table.item(row, column_to_check)
            value = item.text() if item else ""
            should_hide = hide_non_empty and value.strip() != ""
            self.table.setRowHidden(row, should_hide)

    def evaluate(self):
        """Evaluate responses using AI and display confidence thresholds."""
        df_result = evaluate_dataframe(self.df, test=False)
        self.df = df_result
        self.write_table(resize=False)

        # Calculate confidence thresholds
        confidence_series = pd.to_numeric(df_result['confidence'], errors='coerce').dropna()
        total_valid = len(confidence_series)
        umbrales = sorted(confidence_series.unique(), reverse=True)
        
        # Prepare threshold data for display
        data = []
        for umbral in umbrales:
            count = (confidence_series >= umbral).sum()
            porcentaje = (count / total_valid) * 100
            data.append((umbral, porcentaje))

        # Populate thresholds table
        self.umbrales.setRowCount(len(data))
        self.umbrales.setColumnCount(2)
        self.umbrales.setHorizontalHeaderLabels(['Umbral', 'Porcentaje'])

        for row, (umbral, porcentaje) in enumerate(data):
            umbral_item = QTableWidgetItem(f"{umbral:.2f}")
            porcentaje_item = QTableWidgetItem(f"{porcentaje:.2f}%")
            umbral_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            porcentaje_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            self.umbrales.setItem(row, 0, umbral_item)
            self.umbrales.setItem(row, 1, porcentaje_item)
        
        return

    def _dataframe_generation_from_table(self, table):
        """
        Create a DataFrame from the current table contents.

        Args:
            table: The QTableWidget to convert

        Returns:
            DataFrame containing the table data
        """
        number_of_rows = table.rowCount()
        number_of_columns = table.columnCount()

        tmp_df = pd.DataFrame( 
                    columns=['Respuesta', 'Nota', 'Freq', '%'], 
                    index=range(number_of_rows) 
                    ) 
        
        for i in range(number_of_rows):
            for j in range(number_of_columns):
                tmp_df.iloc[i, j] = str(table.item(i, j).text())  

        return tmp_df   
    
    def eventFilter(self, source, event):
        """
        Handle mouse events for quick grading in the table.

        Args:
            source: The source of the event
            event: The event object

        Returns:
            bool: True if event was handled, False otherwise
        """
        
        if source == self.table.viewport() and isinstance(event, QMouseEvent):
            item = self.table.itemAt(event.pos())
            
            if item is not None and self.radioEdicionNotas.isChecked():
                try:
                    nota_col = list(self.df.columns).index("Nota")
                    if item.column() == nota_col:
                        if event.button() == Qt.MouseButton.LeftButton:
                            item.setText("1")
                            return True
                        elif event.button() == Qt.MouseButton.RightButton:
                            item.setText("0")
                            return True
                        elif event.button() == Qt.MouseButton.MiddleButton:
                            item.setText("")
                            return True
                except ValueError:
                    pass
        
        return super().eventFilter(source, event)

    def _close_and_save(self):
        """Save AI evaluations to Excel and close the application."""
        umbral_final = float(self.umbral.toPlainText())

        for row in range(self.table.rowCount()):
            score_AI = self.table.item(row, 4).text()
            feedback = self.table.item(row, 5).text()
            confidence = self.table.item(row, 6).text()
            text = self.table.item(row, 0).text()

            if confidence != '':
                if float(confidence) < umbral_final:
                    continue

            if text not in self.relations:
                self.dlg.setText("Error en el texto: ", text, " -> ESTE ERROR NO DEBERIA PASSAR")
                button = self.dlg.exec()
                if button == QMessageBox.StandardButton.Ok:
                    return
            

            self.excel.cell(row=1, column=self.idxCorr + 2).value = "Nota IA"
            self.excel.cell(row=1, column=self.idxCorr + 3).value = "Feedback IA"
            self.excel.cell(row=1, column=self.idxCorr + 4).value = "Confidence"
            # Write full evaluation to Excel
            for idRow in self.relations[text]:
                # Write AI score
                cell_to_write = self.excel.cell(row=idRow + 2, column=self.idxCorr + 2)
                cell_to_write.value = str(score_AI)

                # Write feedback
                cell_to_write = self.excel.cell(row=idRow + 2, column=self.idxCorr + 3)
                cell_to_write.value = str(feedback)

                # Write confidence
                cell_to_write = self.excel.cell(row=idRow + 2, column=self.idxCorr + 4)
                cell_to_write.value = str(confidence)

        # Save to outputs directory
        self.wb.save(os.getcwd() + '\outputs\\' + self.path.name.replace(".xlsx"," - corregido") + ".xlsx")

        sys.exit()